from odoo import _, models, fields, api
from odoo.exceptions import UserError, ValidationError
import requests
import base64
import io
import re
import xlsxwriter
from openpyxl import load_workbook
from PIL import Image
from io import BytesIO

class ImportContactWizard(models.TransientModel):
    _name = 'import.contact.wizard'
    _description = 'Import Contact Wizard'

    file = fields.Binary(string="Archivo", required=False)

    def is_valid_image_base64(self, b64_string):
        try:
            decoded = base64.b64decode(b64_string, validate=True)
            Image.open(BytesIO(decoded)).verify()
            return True
        except Exception:
            return False
    
    def is_valid_base64(self, string):
        try:
            cleaned = re.sub(r'[^A-Za-z0-9+/=]', '', str(string or ''))
            padding = len(cleaned) % 4
            if padding:
                cleaned += '=' * (4 - padding)
            base64.b64decode(cleaned)
            return True
        except Exception:
            return False   

    def get_image_from_drive(self, shared_url):
        import logging
        _logger = logging.getLogger(__name__)
        try:
            session = requests.Session()
            response = session.get(shared_url, stream=True, timeout=15)

            if "Content-Disposition" not in response.headers:
                file_id_match = re.search(r'/d/([a-zA-Z0-9_-]+)', shared_url)
                if not file_id_match:
                    _logger.warning(f"[Drive] Invalid shared link format: {shared_url}")
                    return None
                file_id = file_id_match.group(1)

                confirm_token = None
                for key, value in response.cookies.items():
                    if key.startswith('download_warning'):
                        confirm_token = value
                        break

                download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
                if confirm_token:
                    download_url += f"&confirm={confirm_token}"

                response = session.get(download_url, stream=True, timeout=15)

            if response.status_code != 200:
                _logger.warning(f"[Drive] Failed to fetch image. Status: {response.status_code}")
                return None

            content_type = response.headers.get('Content-Type', '')
            if 'image' not in content_type.lower():
                _logger.warning(f"[Drive] Not an image: Content-Type: {content_type}")
                return None

            img = Image.open(BytesIO(response.content)).convert("RGB")
            buffer = BytesIO()
            img.save(buffer, format='PNG')
            return base64.b64encode(buffer.getvalue()).decode('utf-8')

        except Exception as e:
            _logger.warning(f"[Drive] Error downloading image from shared link: {e}")
            return None

    def action_test(self):
        if not self.file:
            raise UserError(_("You must upload an XLSX file."))

        try:
            file_content = base64.b64decode(self.file)
            workbook = load_workbook(io.BytesIO(file_content))
        except Exception:
            raise UserError(_("Failed to read the file. Make sure it's a valid XLSX format."))

        sheet = workbook.active
        data = [row for row in sheet.iter_rows(min_row=2, values_only=True)]

        validation_errors = self.validate_data(data)
        if validation_errors:
            formatted_errors = "\n".join(f"• {err}" for err in validation_errors)
            raise UserError(_("Errors found in the records:\n\n") + formatted_errors)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Validation successful'),
                'message': _('Everything looks good in the uploaded file!'),
                'type': 'success',
                'sticky': False,
            }
        }

    def validate_data(self, data):
        errors = []
        seen_names = set()
        seen_trade_names = set()

        existing_tags = set(self.env['res.partner.category'].search([]).mapped('name'))
        existing_accounts = set(self.env['account.account'].search([]).mapped('name'))
        existing_terms = set(self.env['account.payment.term'].search([]).mapped('name'))
        existing_pricelists = set(self.env['product.pricelist'].search([]).mapped('name'))
        existing_states = set(self.env['res.country.state'].search([]).mapped('name'))

        for index, row in enumerate(data[1:], start=3):
            if len(row) < 22:
                errors.append(f"Row {index}: Not enough columns (expected 22).")
                continue

            (
                is_company_str, is_customer_str, is_supplier_str, name, trade_name,
                id_type, identifier,
                state_name, city, street, reference,
                phone, mobile, email,
                website, tags_str, account_receivable, account_payable,
                payment_term, pricelist, electronic_docs_str, image_input
            ) = row[:22]

            for value, label in [
                (is_company_str, 'Is Company'),
                (is_customer_str, 'Is Customer'),
                (is_supplier_str, 'Is Supplier'),
                (electronic_docs_str, 'Electronic Documents')
            ]:
                if str(value).strip().lower() not in ['true', 'false']:
                    errors.append(f"Row {index}: Field '{label}' must be 'True' or 'False'.")

            if not name or not str(name).strip():
                errors.append(f"Row {index}: Field 'Name' cannot be empty.")
            else:
                name_key = str(name).strip().lower()
                if name_key in seen_names:
                    errors.append(f"Row {index}: Duplicate Name '{name}'.")
                seen_names.add(name_key)

            if trade_name and str(trade_name).strip():
                trade_key = str(trade_name).strip().lower()
                if trade_key in seen_trade_names:
                    errors.append(f"Row {index}: Duplicate Trade Name '{trade_name}'.")
                seen_trade_names.add(trade_key)

            id_type_clean = str(id_type or '').strip().upper()
            if id_type_clean not in ['CEDULA', 'RUC', 'PASAPORTE']:
                errors.append(f"Row {index}: ID Type must be 'CEDULA', 'RUC' or 'PASAPORTE'.")

            identifier_str = str(identifier or '').strip()
            if not identifier_str.isdigit() or not (10 <= len(identifier_str) <= 13):
                errors.append(f"Row {index}: Identifier must be a numeric value between 10 and 13 digits.")

            if not state_name or str(state_name).strip() not in existing_states:
                errors.append(f"Row {index}: State '{state_name}' is invalid or not registered.")
            if not city or not str(city).strip():
                errors.append(f"Row {index}: City is required.")
            if not street or not str(street).strip():
                errors.append(f"Row {index}: Street (Address 1) is required.")
            if not reference or not str(reference).strip():
                errors.append(f"Row {index}: Reference (Address 2) is required.")

            phone_str = str(phone or '').strip()
            if phone_str:
                phone_clean = phone_str.replace(" ", "")
                if not re.match(r'^(02|03|04|05|06|07)\s?\d{7}$', phone_str) and not re.match(r'^(\+593|09)\d{7,17}$', phone_clean):
                    errors.append(f"Row {index}: Invalid phone format. Use '+593...', '09...', or '02 2345678'.")

            mobile_str = str(mobile or '').strip()
            if mobile_str:
                mobile_clean = mobile_str.replace(" ", "")
                if not (mobile_clean.startswith('+593') or mobile_clean.startswith('09') or mobile_clean.startswith('02250')):
                    errors.append(f"Row {index}: Mobile must start with '+593', '09', or '02250'.")
                elif not (10 <= len(mobile_clean) <= 20):
                    errors.append(f"Row {index}: Mobile length must be 10–20 digits.")

            if not email or not str(email).strip():
                errors.append(f"Row {index}: Email is required.")

            if tags_str:
                for tag in [t.strip() for t in str(tags_str).split(',') if t.strip()]:
                    if tag not in existing_tags:
                        errors.append(f"Row {index}: Tag '{tag}' does not exist.")

            if account_receivable and str(account_receivable).strip() not in existing_accounts:
                errors.append(f"Row {index}: Account Receivable '{account_receivable}' does not exist.")
            if account_payable and str(account_payable).strip() not in existing_accounts:
                errors.append(f"Row {index}: Account Payable '{account_payable}' does not exist.")

            if payment_term and str(payment_term).strip() not in existing_terms:
                errors.append(f"Row {index}: Payment Term '{payment_term}' does not exist.")
            if pricelist and str(pricelist).strip() not in existing_pricelists:
                errors.append(f"Row {index}: Pricelist '{pricelist}' does not exist.")

            if image_input:
                is_valid = False

                if self.is_valid_base64(image_input):
                    is_valid = True
                elif isinstance(image_input, str) and 'drive.google.com/file/d/' in image_input:
                    is_valid = True
                elif self.get_image_from_drive(image_input):
                    is_valid = True

                if not is_valid:
                    errors.append(f"Row {index}: Invalid image format or inaccessible Drive link.")

        return errors

    def action_export_template(self):
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})

        header_format = workbook.add_format({
            'bold': True, 'text_wrap': True, 'valign': 'vcenter',
            'align': 'center', 'border': 1, 'bg_color': '#D9E1F2'
        })
        example_format = workbook.add_format({
            'font_color': '#0070C0', 'valign': 'vcenter', 'align': 'left', 'border': 1
        })
        header_notes_format = workbook.add_format({
            'bold': True, 'text_wrap': True, 'valign': 'vcenter',
            'align': 'center', 'border': 1, 'bg_color': '#4CAF50', 'color': 'white'
        })
        notes_format = workbook.add_format({
            'text_wrap': True, 'valign': 'top', 'align': 'left', 'border': 1, 'bg_color': '#F9F9F9'
        })

        headers = [
            'Es Empresa', 'Es Cliente', 'Es Proveedor', 'Nombre', 'Nombre Comercial',
            'Tipo de Identificación', 'Identificación',
            'Provincia', 'Ciudad', 'Calle (Dirección 1)', 'Referencia (Dirección 2)',
            'Teléfono', 'Celular', 'Correo Electrónico',
            'Sitio Web', 'Etiquetas', 'Cuenta por Cobrar', 'Cuenta por Pagar',
            'Plazo de Pago', 'Lista de Precios', 'Documentos Electrónicos', 'Imagen'
        ]

        example_data = [
           'True', 'True', 'False', 'Ejemplo S.A.', 'Comercial Ejemplo',
            'CEDULA', '0912345678',
            'Pichincha', 'Quito', 'Av. Amazonas', 'Edificio ABC',
            '+59322345678', '+593998765432', 'info@ejemplo.com',
            'https://ejemplo.com', 'Importador,Cliente VIP', 'ANTICIPOS EMPLEADO', 'ANTICIPOS CLIENTE',
            'Pago Inmediato', 'Lista de Precios Estándar', 'True',
            'https://drive.google.com/file/d/1ABCDEF23456789GHIJKL/view?usp=sharing'
        ]

        notes = [
            'Es Empresa: "True" si el contacto es una empresa, "False" en caso contrario.',
            'Es Cliente: "True" si el contacto es un cliente.',
            'Es Proveedor: "True" si el contacto es un proveedor.',
            'Nombre: Obligatorio. Nombre principal del contacto o empresa.',
            'Nombre Comercial: Opcional. Nombre alternativo.',
            'Tipo de Identificación: Debe ser "CEDULA", "RUC" o "PASAPORTE".',
            'Identificación: Número de 10 a 13 dígitos. Solo se permiten números.',
            'Provincia: Debe coincidir con un nombre de provincia registrado.',
            'Ciudad: Obligatorio. Ciudad del contacto.',
            'Calle (Dirección 1): Calle o avenida principal.',
            'Referencia (Dirección 2): Referencia adicional (ej. edificio).',
            'Teléfono: Opcional. Usa formatos como "+593...", "09..." o fijos como "02 2345678".',
            'Celular: Debe comenzar con "+593" o "09".',
            'Correo Electrónico: Obligatorio. Debe ser válido.',
            'Sitio Web: Opcional. Debe ser una URL válida.',
            'Etiquetas: Opcional. Nombres separados por comas (deben existir en el sistema).',
            'Cuenta por Cobrar: Debe coincidir con un nombre de cuenta existente.',
            'Cuenta por Pagar: Debe coincidir con un nombre de cuenta existente.',
            'Plazo de Pago: Debe coincidir con un nombre de término de pago existente.',
            'Lista de Precios: Debe coincidir con un nombre de lista de precios existente.',
            'Documentos Electrónicos: "True" para habilitar la facturación electrónica.',
            'Imagen: Debes proporcionar un enlace compartido a una imagen almacenada en Google Drive. El archivo debe estar compartido con la opción "Cualquiera con el enlace - Lector" y el enlace debe ser copiado usando la opción "Copiar enlace" desde Google Drive. Ejemplo: https://drive.google.com/file/d/FILE_ID/view?usp=sharing'
        ]

        sheet = workbook.add_worksheet('Plantilla Contacto')
        sheet.set_column(0, len(headers) - 1, 25)
        for col, header in enumerate(headers):
            sheet.write(0, col, header, header_format)
        for col, example in enumerate(example_data):
            sheet.write(1, col, example, example_format)

        sheet_notes = workbook.add_worksheet('Observaciones')
        sheet_notes.set_column('A:A', 90)
        sheet_notes.write('A1', 'Observaciones', header_notes_format)
        for row, note in enumerate(notes, start=1):
            sheet_notes.write(row, 0, note, notes_format)

        workbook.close()
        file_data = base64.b64encode(output.getvalue())
        output.close()

        attachment = self.env['ir.attachment'].create({
            'name': 'contact_template.xlsx',
            'type': 'binary',
            'datas': file_data,
            'res_model': 'import.contact.wizard',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_import(self):
        if not self.file:
            raise UserError("You must upload an XLSX file.")

        try:
            file_content = base64.b64decode(self.file)
            workbook = load_workbook(io.BytesIO(file_content))
        except Exception:
            raise UserError("Failed to read the file. Please make sure it is a valid XLSX format.")

        sheet = workbook.active
        data = [row for row in sheet.iter_rows(min_row=2, values_only=True)]

        errors = []
        updated_contacts = []
        created_count = 0

        partner_model = self.env['res.partner']
        country = self.env['res.country'].search([('code', '=', 'EC')], limit=1)

        for index, row in enumerate(data[1:], start=3):
            try:
                if len(row) < 22:
                    errors.append(f"Row {index}: Not enough columns.")
                    continue

                (
                    is_company_str, is_customer_str, is_supplier_str, name, trade_name,
                    id_type, identifier,
                    state_name, city, street, reference,
                    phone, mobile, email,
                    website, tags_str, account_receivable, account_payable,
                    payment_term, pricelist, electronic_docs_str, image_input
                ) = row

                id_type_clean = str(id_type or '').strip().upper()
                identifier_clean = str(identifier or '').strip()
                type_identifier = {'CEDULA': 'cedula', 'RUC': 'ruc', 'PASAPORTE': 'passport'}.get(id_type_clean)

                if not type_identifier:
                    errors.append(f"Row {index}: Invalid Identification Type '{id_type_clean}'.")
                    continue

                is_company = str(is_company_str).strip().lower() == 'true'
                is_customer = str(is_customer_str).strip().lower() == 'true'
                is_supplier = str(is_supplier_str).strip().lower() == 'true'
                is_electronic = str(electronic_docs_str).strip().lower() == 'true'

                existing_contact = partner_model.search([
                    ('identifier', '=', identifier_clean),
                    ('type_identifier', '=', type_identifier),
                ], limit=1)

                state = self.env['res.country.state'].search([
                    ('name', 'ilike', state_name),
                    ('country_id', '=', country.id)
                ], limit=1)

                tag_ids = []
                if tags_str:
                    tag_names = [tag.strip() for tag in tags_str.split(',') if tag.strip()]
                    tag_ids = self.env['res.partner.category'].search([('name', 'in', tag_names)]).ids

                payment_term_id = self.env['account.payment.term'].search([('name', 'ilike', payment_term)], limit=1)
                pricelist_id = self.env['product.pricelist'].search([('name', 'ilike', pricelist)], limit=1)
                account_recv_id = self.env['account.account'].search([('name', 'ilike', account_receivable or 'EMPLOYEE ADVANCES')], limit=1)
                account_pay_id = self.env['account.account'].search([('name', 'ilike', account_payable or 'CUSTOMER ADVANCES')], limit=1)

                image_bytes = False
                if image_input:
                    if self.is_valid_base64(image_input):
                        image_bytes = image_input
                    elif isinstance(image_input, str) and 'drive.google.com/file/d/' in image_input:
                        downloaded = self.get_image_from_drive(image_input)
                        if downloaded:
                            image_bytes = downloaded
                        else:
                            errors.append(f"Row {index}: Could not download image from Google Drive link.")
                    else:
                        downloaded = self.get_image_from_drive(image_input)
                        if downloaded:
                            image_bytes = downloaded
                        else:
                            errors.append(f"Row {index}: Invalid image format or failed download: {image_input}")

                values = {
                    'name': name,
                    'trade_name': trade_name,
                    'is_company': is_company,
                    'phone': phone,
                    'mobile': mobile,
                    'email': email,
                    'website': website,
                    'category_id': [(6, 0, tag_ids)] if tag_ids else False,
                    'customer_rank': 1 if is_customer else 0,
                    'supplier_rank': 1 if is_supplier else 0,
                    'identifier': identifier_clean,
                    'type_identifier': type_identifier,
                    'is_electronic': is_electronic,
                    'street': street or '',
                    'street2': reference or '',
                    'city': city or '',
                    'state_id': state.id if state else False,
                    'country_id': country.id if country else False,
                    'property_payment_term_id': payment_term_id.id if payment_term_id else False,
                    'property_product_pricelist': pricelist_id.id if pricelist_id else False,
                    'property_account_receivable_id': account_recv_id.id if account_recv_id else False,
                    'property_account_payable_id': account_pay_id.id if account_pay_id else False,
                    'image_1920': image_bytes or False,
                }

                if existing_contact:
                    existing_contact.write(values)
                    updated_contacts.append(existing_contact.name)
                else:
                    partner_model.create(values)
                    created_count += 1

            except Exception as e:
                errors.append(f"Row {index}: Failed to create or update contact: {str(e)}")

        if errors:
            formatted = "\n".join(f"• {e}" for e in errors)
            raise UserError("Errors during import:\n\n" + formatted)

        messages = []
        if created_count:
            messages.append(f"{created_count} new contact(s) created.")
        if updated_contacts:
            updated_names = "\n".join(f"• {name}" for name in updated_contacts)
            messages.append(f"{len(updated_contacts)} contact(s) updated:\n{updated_names}")

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Import Results',
                'message': "\n\n".join(messages) if messages else "No changes were made.",
                'type': 'success',
                'sticky': False,
            }
        }