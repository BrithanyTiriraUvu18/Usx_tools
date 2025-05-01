from odoo import _, models, fields, api
from odoo.exceptions import UserError
import base64
import xlsxwriter
from io import BytesIO
import io
import csv
import openpyxl


class WizardPricelist(models.TransientModel):
    _name = "wizard.pricelist"
    _description = "Importar Lista de Precios"

    file = fields.Binary(string="Archivo", required=True)
    file_name = fields.Char(string="Nombre del Archivo")
    pricelist_id = fields.Many2one('product.pricelist', string="Lista de precios")
    product_ids = fields.Many2many('product.template', string="Productos")

    def export_template(self):
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})

        header_format = workbook.add_format({
            'bold': True, 'text_wrap': True, 'valign': 'vcenter',
            'align': 'center', 'border': 1, 'bg_color': '#D9E1F2'
        })
        observations_format = workbook.add_format({
            'bold': False, 'text_wrap': True, 'valign': 'top',
            'align': 'left', 'border': 1, 'bg_color': '#F9F9F9'
        })

        worksheet = workbook.add_worksheet('Plantilla Lista de Precios')
        headers = ['Referencia Interna', 'Nombre de la lista de precios', 'Precio']
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)
        worksheet.set_column(0, len(headers) - 1, 25)

        row = 1
        for product in self.product_ids:
            pricelist_item = self.env['product.pricelist.item'].search([
                ('product_tmpl_id', '=', product.id),
                ('pricelist_id', '=', self.pricelist_id.id)
            ])
            if pricelist_item:
                worksheet.write(row, 0, product.default_code)
                worksheet.write(row, 1, self.pricelist_id.name)
                worksheet.write(row, 2, pricelist_item.fixed_price)
                row += 1

        worksheet2 = workbook.add_worksheet('Observaciones')
        worksheet2.write('A1', 'Observaciones', header_format)
        observaciones = [
            'Referencia Interna: El código único del producto.',
            'Nombre de la lista de precios: Por ejemplo, "Mayorista", "Distribuidor", etc.',
            'Precio: El valor numérico del producto en dicha lista de precios.'
        ]
        for row_num, obs in enumerate(observaciones, start=2):
            worksheet2.write(row_num, 0, obs, observations_format)
        worksheet2.set_column('A:A', 80)

        workbook.close()
        file_data = base64.b64encode(output.getvalue())
        output.close()

        attachment = self.env['ir.attachment'].create({
            'name': 'plantilla_lista_de_precios.xlsx',
            'type': 'binary',
            'datas': file_data,
            'res_model': 'wizard.pricelist',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
    
    def test_data(self):
        if not self.file:
            raise UserError("No has subido ningún archivo.")

        file_name = self.file_name or 'archivo.xlsx'
        extension = file_name.lower().split('.')[-1]
        if extension not in ['xlsx', 'csv']:
            raise UserError("Solo se permiten archivos con extensión .xlsx o .csv.")

        data = base64.b64decode(self.file)
        records = self._leer_datos(extension, data)
        self._validar_unicidad(records)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Validación Exitosa',
                'message': f'Se validaron {len(records)} registros correctamente.',
                'type': 'success',
                'sticky': False,
            }
        }

    def _leer_datos(self, extension, data):
        records = []
        if extension == 'xlsx':
            workbook = openpyxl.load_workbook(filename=io.BytesIO(data), data_only=True)
            sheet = workbook.active
            filas = ((i, row) for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2))
        else:
            decoded = data.decode('utf-8')
            reader = csv.reader(io.StringIO(decoded))
            next(reader, None)
            filas = ((i, row) for i, row in enumerate(reader, start=2))

        for i, row in filas:
            if not row or len(row) < 3 or all((str(cell).strip() == '' if cell is not None else True) for cell in row):
                continue
            record = self._procesar_fila(row, i)
            records.append(record)

        return records

    def _procesar_fila(self, row, fila_num):
        try:
            ref, lista, precio = row[:3]
        except ValueError:
            raise UserError(f"Fila {fila_num}: Debe tener exactamente 3 columnas.")

        if not ref or not lista or precio in (None, ''):
            raise UserError(f"Fila {fila_num}: Todos los campos son obligatorios.")

        return {
            'default_code': str(ref).strip(),
            'pricelist_name': str(lista).strip(),
            'price': precio
        }

    def _validar_unicidad(self, records):
        referencia_map = {}
        for rec in records:
            ref = rec['default_code']
            lista = rec['pricelist_name']
            if ref not in referencia_map:
                referencia_map[ref] = lista
            elif referencia_map[ref] != lista:
                raise UserError(
                    f"La referencia '{ref}' está asociada a múltiples listas de precios ('{referencia_map[ref]}' y '{lista}')."
                )

    def import_data(self):
        if not self.file:
            raise UserError("No has subido ningún archivo.")

        file_name = (self.file_name or 'archivo.xlsx').strip()
        extension = file_name.lower().rsplit('.', 1)[-1]
        if extension not in ['xlsx', 'csv']:
            raise UserError("Solo se permiten archivos con extensión .xlsx o .csv.")

        data = base64.b64decode(self.file)
        records = self._leer_datos(extension, data)

        creados = 0
        actualizados = 0
        errores = []

        for i, rec in enumerate(records, start=2):
            ref = rec['default_code']
            lista_nombre = rec['pricelist_name']
            precio = rec['price']
             
            #Busca en Odoo un producto cuyo código interno sea igual al valor de esta fila del Excel.
            producto = self.env['product.template'].search([('default_code', '=', ref)], limit=1)
            if not producto:
                errores.append(f"Fila {i}: Producto con referencia '{ref}' no encontrado.")
                continue
            #Busca en Odoo una lista de precios cuyo nombre sea igual al valor de esta fila del Excel.
            lista_precio = self.env['product.pricelist'].search([('name', '=', lista_nombre)], limit=1)
            if not lista_precio:
                errores.append(f"Fila {i}: Lista de precios '{lista_nombre}' no encontrada.")
                continue

            item = self.env['product.pricelist.item'].search([
                ('product_tmpl_id', '=', producto.id),
                ('pricelist_id', '=', lista_precio.id)
            ], limit=1)

            vals = {
                'product_tmpl_id': producto.id,
                'pricelist_id': lista_precio.id,
                'fixed_price': precio,
                'applied_on': '1_product',
                'compute_price': 'fixed'
            }

            if item:
                item.write({'fixed_price': precio})
                actualizados += 1
            else:
                self.env['product.pricelist.item'].create(vals)
                creados += 1

        # Mostrar notificación
        mensaje = f"Importación completada.\nCreados: {creados}, Actualizados: {actualizados}"
        if errores:
            mensaje += "\n\nErrores:\n" + "\n".join(f"- {e}" for e in errores)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Resultado de la Importación',
                'message': mensaje,
                'type': 'warning' if errores else 'success',
                'sticky': False,
            }
        }
